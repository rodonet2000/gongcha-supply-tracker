#!/usr/bin/env python3
"""
Import Medidas por Bebida from Conteo_Quincenal_Liz.xlsx into gongcha.menu_items + gongcha.recipes.

Expected Excel structure (Sheet "Medidas por Bebida"):
  - Row 1 (header): Bebida | Insumo | [Mediano] | [Grande] | Unidad | Notas
  OR a matrix layout where column A = bebida, column B = insumo, col C = qty_mediano, col D = qty_grande

Run first with --inspect to see actual structure before importing.
Usage:
  python scripts/import_medidas.py --inspect          # show Excel structure
  python scripts/import_medidas.py --dry-run          # generate SQL without executing
  python scripts/import_medidas.py                    # import to DB
"""
import sys
import os
import paramiko

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Conteo_Quincenal_Liz.xlsx')
VPS_HOST   = '5.252.53.169'
VPS_USER   = 'root'
VPS_PASS   = 'Rodonet7012'
DB         = 'supabase-db-h8occ6uko144qwdes4o43t7r'
SHEET_NAME = 'Medidas por Bebida'


def inspect_excel():
    """Print the first 10 rows with column headers to understand the layout."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Available sheets: {wb.sheetnames}\n")

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    print(f"Sheet dimensions: {ws.dimensions}")
    print(f"\nFirst 15 rows (col A–H):\n")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        print(f"  Row {row_idx:2d}: {[str(v)[:30] if v else '' for v in row[:8]]}")


def parse_recipes():
    """
    Parse the Medidas por Bebida sheet.

    Adjust COL_* constants below to match the actual layout from --inspect.
    Current assumption (common Gon-Cha format):
      Col A = Bebida name (merged across ingredient rows)
      Col B = Insumo name
      Col C = Cantidad Mediano
      Col D = Cantidad Grande
      Col E = Unidad (if present)
    Returns list of (bebida, insumo, size, cantidad, unidad) tuples.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # ──────────────────────────────────────────────────────
    # ADJUST THESE COLUMN INDICES (0-based) to match layout:
    COL_BEBIDA   = 0   # Column A: drink name
    COL_INSUMO   = 1   # Column B: ingredient name
    COL_MEDIANO  = 2   # Column C: quantity for Mediano
    COL_GRANDE   = 3   # Column D: quantity for Grande
    COL_UNIDAD   = 4   # Column E: unit (optional)
    HEADER_ROWS  = 1   # number of rows to skip at top
    # ──────────────────────────────────────────────────────

    records = []
    current_bebida = None

    for row in ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
        # Track merged bebida cell (value only appears in first row of merge)
        if row[COL_BEBIDA] and str(row[COL_BEBIDA]).strip():
            current_bebida = str(row[COL_BEBIDA]).strip()

        if not current_bebida:
            continue

        insumo = row[COL_INSUMO]
        if not insumo or not str(insumo).strip():
            continue

        insumo = str(insumo).strip()
        unidad = str(row[COL_UNIDAD]).strip() if COL_UNIDAD < len(row) and row[COL_UNIDAD] else None

        qty_med = row[COL_MEDIANO]
        qty_gde = row[COL_GRANDE]

        if qty_med and str(qty_med).strip() not in ('', '-', 'N/A', 'n/a'):
            try:
                records.append((current_bebida, insumo, 'Mediano', float(qty_med), unidad))
            except (ValueError, TypeError):
                pass

        if qty_gde and str(qty_gde).strip() not in ('', '-', 'N/A', 'n/a'):
            try:
                records.append((current_bebida, insumo, 'Grande', float(qty_gde), unidad))
            except (ValueError, TypeError):
                pass

    return records


def build_sql(records):
    """Generate the SQL to insert menu_items and recipes."""
    bebidas = sorted(set(r[0] for r in records))
    insumos = sorted(set(r[1] for r in records))

    lines = ["BEGIN;", ""]
    lines.append("-- Insert menu items (drinks)")
    for b in bebidas:
        safe = b.replace("'", "''")
        lines.append(
            f"INSERT INTO gongcha.menu_items (name, category) VALUES ('{safe}', 'bebida') "
            f"ON CONFLICT (name) DO NOTHING;"
        )

    lines.append("")
    lines.append("-- Insert recipes (join via subquery to get IDs)")
    for bebida, insumo, size, cantidad, _ in records:
        sb = bebida.replace("'", "''")
        si = insumo.replace("'", "''")
        lines.append(
            f"INSERT INTO gongcha.recipes (menu_item_id, supply_id, quantity, size) "
            f"SELECT m.id, s.id, {cantidad}, '{size}' "
            f"FROM gongcha.menu_items m, gongcha.supplies s "
            f"WHERE m.name = '{sb}' AND s.name = '{si}' "
            f"ON CONFLICT (menu_item_id, supply_id, size) DO UPDATE SET quantity = EXCLUDED.quantity;"
        )

    lines.append("")
    lines.append("COMMIT;")
    return "\n".join(lines)


def run_sql(sql):
    """Execute SQL on the remote DB via SSH."""
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(VPS_HOST, username=VPS_USER, password=VPS_PASS, timeout=20)

    cmd = f'docker exec -i {DB} psql -U postgres -d postgres'
    stdin, stdout, stderr = ssh.exec_command(cmd, timeout=120)
    stdin.write(sql.encode())
    stdin.channel.shutdown_write()

    out = stdout.read().decode('utf-8', errors='replace')
    err = stderr.read().decode('utf-8', errors='replace')
    rc  = stdout.channel.recv_exit_status()
    ssh.close()
    return out, err, rc


def main():
    dry_run = '--dry-run' in sys.argv
    inspect = '--inspect' in sys.argv

    if inspect:
        inspect_excel()
        return

    records = parse_recipes()
    if not records:
        print("No records parsed. Run --inspect first to verify the sheet layout.")
        sys.exit(1)

    print(f"Parsed {len(records)} recipe entries across {len(set(r[0] for r in records))} drinks.\n")

    # Report unmatched supplies (names not found in DB would silently be skipped by the SQL)
    unique_insumos = sorted(set(r[1] for r in records))
    print(f"Distinct insumos referenced: {len(unique_insumos)}")
    for name in unique_insumos[:10]:
        print(f"  - {name}")
    if len(unique_insumos) > 10:
        print(f"  ... and {len(unique_insumos) - 10} more")

    sql = build_sql(records)

    if dry_run:
        print("\n── DRY RUN SQL (first 40 lines) ──")
        for line in sql.split('\n')[:40]:
            print(line)
        print("...")
        return

    print("\nExecuting on remote DB...")
    out, err, rc = run_sql(sql)
    print(out)
    if err:
        print("STDERR:", err)
    print(f"\nExit code: {rc}")
    if rc == 0:
        print(f"✓ Import complete — {len(set(r[0] for r in records))} drinks, {len(records)} recipe lines")
    else:
        print("✗ Import FAILED — check STDERR above")


if __name__ == '__main__':
    main()
