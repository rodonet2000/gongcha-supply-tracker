#!/usr/bin/env python3
"""
Import Rendimiento por Unidad from Conteo_Quincenal_Liz.xlsx into gongcha.yield_factors.

Expected Excel structure (Sheet "Rendimiento por Unidad"):
  Col A = Insumo (supply name — must match gongcha.supplies.name)
  Col B = Rendimiento (decimal 0–1, e.g. 0.95 = 95%, OR percentage e.g. 95)
  Col C = Descripción (unit context, optional)
  Col D = Notas (optional)

Run with --inspect first to verify column layout.
Usage:
  python scripts/import_rendimiento.py --inspect
  python scripts/import_rendimiento.py --dry-run
  python scripts/import_rendimiento.py
"""
import sys
import os
import paramiko

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Conteo_Quincenal_Liz.xlsx')
VPS_HOST   = '5.252.53.169'
VPS_USER   = 'root'
VPS_PASS   = 'Rodonet7012'
DB         = 'supabase-db-h8occ6uko144qwdes4o43t7r'
SHEET_NAME = 'Rendimiento por Unidad'


def inspect_excel():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Available sheets: {wb.sheetnames}\n")

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    print(f"Sheet dimensions: {ws.dimensions}")
    print(f"\nFirst 20 rows (col A–F):\n")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        print(f"  Row {row_idx:2d}: {[str(v)[:35] if v else '' for v in row[:6]]}")


def parse_yields():
    """
    Parse yield factors from the sheet.

    Adjust COL_* constants to match actual layout from --inspect.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # ──────────────────────────────────────────────────────
    # ADJUST THESE to match actual column layout:
    COL_INSUMO = 0   # Column A
    COL_FACTOR = 1   # Column B  (0.95 or 95 — auto-detected)
    COL_DESC   = 2   # Column C  (optional description)
    COL_NOTES  = 3   # Column D  (optional notes)
    HEADER_ROWS = 1  # rows to skip
    # ──────────────────────────────────────────────────────

    records = []
    for row in ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
        insumo = row[COL_INSUMO] if COL_INSUMO < len(row) else None
        factor = row[COL_FACTOR] if COL_FACTOR < len(row) else None

        if not insumo or not str(insumo).strip():
            continue
        if factor is None or str(factor).strip() in ('', '-'):
            continue

        insumo = str(insumo).strip()
        try:
            f = float(str(factor).replace('%', '').strip())
        except ValueError:
            continue

        # Normalize: if given as percentage (>1), convert to decimal
        if f > 1:
            f = f / 100.0

        if not (0 < f <= 1):
            print(f"  SKIP: '{insumo}' has invalid factor {f} (must be 0–1)")
            continue

        desc  = str(row[COL_DESC]).strip() if COL_DESC < len(row) and row[COL_DESC] else None
        notes = str(row[COL_NOTES]).strip() if COL_NOTES < len(row) and row[COL_NOTES] else None

        records.append((insumo, round(f, 4), desc, notes))

    return records


def build_sql(records):
    lines = ["BEGIN;", ""]
    lines.append("-- Insert yield factors (join supply by name)")
    for insumo, factor, desc, notes in records:
        si  = insumo.replace("'", "''")
        sd  = desc.replace("'", "''")  if desc  else None
        sn  = notes.replace("'", "''") if notes else None
        desc_val  = f"'{sd}'"  if sd  else 'NULL'
        notes_val = f"'{sn}'"  if sn  else 'NULL'

        lines.append(
            f"INSERT INTO gongcha.yield_factors (supply_id, factor, unit_description, notes) "
            f"SELECT id, {factor}, {desc_val}, {notes_val} FROM gongcha.supplies WHERE name = '{si}' "
            f"ON CONFLICT (supply_id) DO UPDATE SET factor = EXCLUDED.factor, "
            f"unit_description = EXCLUDED.unit_description, notes = EXCLUDED.notes;"
        )

    lines.append("")
    lines.append("COMMIT;")
    return "\n".join(lines)


def run_sql(sql):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(VPS_HOST, username=VPS_USER, password=VPS_PASS, timeout=20)

    cmd = f'docker exec -i {DB} psql -U postgres -d postgres'
    stdin, stdout, stderr = ssh.exec_command(cmd, timeout=60)
    stdin.write(sql.encode())
    stdin.channel.shutdown_write()

    out = stdout.read().decode('utf-8', errors='replace')
    err = stderr.read().decode('utf-8', errors='replace')
    rc  = stdout.channel.recv_exit_status()
    ssh.close()
    return out, err, rc


def main():
    dry_run = '--dry-run' in sys.argv
    inspect = '--inspect' in sys.argv

    if inspect:
        inspect_excel()
        return

    records = parse_yields()
    if not records:
        print("No records parsed. Run --inspect to verify sheet layout and column indices.")
        sys.exit(1)

    print(f"Parsed {len(records)} yield factors:\n")
    for name, factor, desc, _ in records[:15]:
        print(f"  {name:<40} → {factor:.4f} ({factor*100:.1f}%)  [{desc or ''}]")
    if len(records) > 15:
        print(f"  ... and {len(records) - 15} more")

    sql = build_sql(records)

    if dry_run:
        print("\n── DRY RUN SQL (first 20 lines) ──")
        for line in sql.split('\n')[:20]:
            print(line)
        print("...")
        return

    print("\nExecuting on remote DB...")
    out, err, rc = run_sql(sql)
    print(out)
    if err:
        print("STDERR:", err)
    print(f"\nExit code: {rc}")
    if rc == 0:
        print(f"✓ Import complete — {len(records)} yield factors loaded")
    else:
        print("✗ Import FAILED — check STDERR above")


if __name__ == '__main__':
    main()
